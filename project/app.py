import os
import re
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import numpy as np
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import base64
import qrcode

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'face-attendance-secret-key-change-me')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

db = SQLAlchemy(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='teacher')
    # 关联关系
    classes = db.relationship('Class', backref='user', lazy=True)
    students = db.relationship('Student', backref='user', lazy=True)
    attendances = db.relationship('Attendance', backref='user', lazy=True)

class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    grade = db.Column(db.String(50))
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    # 添加用户外键
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    # 关联关系
    students = db.relationship('Student', backref='class_ref', lazy=True)
    attendances = db.relationship('Attendance', backref='class_ref', lazy=True)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    student_id = db.Column(db.String(50), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    face_descriptor = db.Column(db.LargeBinary)
    face_descriptor_512 = db.Column(db.LargeBinary)
    created_at = db.Column(db.DateTime, default=datetime.now)
    # 添加用户外键
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    # 关联关系
    attendances = db.relationship('Attendance', backref='student_ref', lazy=True)

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    check_in_time = db.Column(db.DateTime, default=datetime.now)
    status = db.Column(db.String(20), default='present')
    method = db.Column(db.String(20), default='face')
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class PendingAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    request_time = db.Column(db.DateTime, default=datetime.now)
    status = db.Column(db.String(20), default='pending')
    method = db.Column(db.String(20), default='qrcode')

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    code = db.Column(db.String(50))
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    classes = db.relationship('Class', secondary='course_class', backref='courses')
    attendances = db.relationship('Attendance', backref='course_ref', lazy=True)

course_class = db.Table('course_class',
    db.Column('course_id', db.Integer, db.ForeignKey('course.id'), primary_key=True),
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True)
)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

with app.app_context():
    db.create_all()
    
    # 数据库迁移：为已有数据库添加 face_descriptor_512 列
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    columns = [c['name'] for c in inspector.get_columns('student')]
    if 'face_descriptor_512' not in columns:
        with db.engine.connect() as conn:
            conn.execute(text('ALTER TABLE student ADD COLUMN face_descriptor_512 BLOB'))
            conn.commit()
        print('[DB] 已添加 face_descriptor_512 列')
    
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            name='管理员',
            role='super_admin'
        )
        db.session.add(admin)
        db.session.commit()

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    today = date.today()
    
    class_count = Class.query.filter_by(user_id=user.id).count()
    student_count = Student.query.filter_by(user_id=user.id).count()
    today_attendance = Attendance.query.filter(
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).count()
    
    attendance_rate = 0
    if student_count > 0:
        attendance_rate = round((today_attendance / student_count) * 100, 1)
    
    class_stats = []
    classes = Class.query.filter_by(user_id=user.id).all()
    for cls in classes:
        total = Student.query.filter_by(class_id=cls.id, user_id=user.id).count()
        attended = Attendance.query.filter(
            Attendance.class_id == cls.id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == today
        ).count()
        rate = round((attended / total) * 100, 1) if total > 0 else 0
        class_stats.append({
            'name': cls.name,
            'total': total,
            'attended': attended,
            'rate': rate
        })
    
    return render_template('index.html',
                         class_count=class_count,
                         student_count=student_count,
                         today_attendance=today_attendance,
                         attendance_rate=attendance_rate,
                         class_stats=class_stats,
                         user=user)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            return render_template('login.html', error='请输入用户名和密码')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session['user_id'] = user.id
            session['username'] = user.username
            session['name'] = user.name
            session['role'] = user.role
            if user.role == 'super_admin':
                return redirect(url_for('super_admin_dashboard'))
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='用户名或密码错误')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        name = request.form.get('name', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not username or not name:
            return render_template('register.html', error='请填写完整的注册信息')
        
        if not re.match(r'^[a-zA-Z0-9_\u4e00-\u9fa5]{2,20}$', username):
            return render_template('register.html', error='用户名需为2-20位字母、数字、下划线和中文')
        
        if len(name) > 50:
            return render_template('register.html', error='姓名不能超过50个字符')
        
        if User.query.filter_by(username=username).first():
            return render_template('register.html', error='用户名已存在')
        
        if password != confirm_password:
            return render_template('register.html', error='两次输入的密码不一致')
        
        if len(password) < 6:
            return render_template('register.html', error='密码长度至少为6位')
        
        if len(password) > 100:
            return render_template('register.html', error='密码长度不能超过100位')
        
        new_user = User(
            username=username,
            name=name,
            password_hash=generate_password_hash(password),
            role='teacher'
        )
        db.session.add(new_user)
        db.session.commit()
        
        return render_template('register.html', success='注册成功！请返回登录页面')
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/courses')
def courses():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    courses = Course.query.filter_by(user_id=user.id).all()
    return render_template('courses.html', courses=courses, user=user)

@app.route('/courses/add', methods=['GET', 'POST'])
def add_course():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            return render_template('add_course.html', error='请输入课程名称', user=user)
        
        if len(name) > 100:
            return render_template('add_course.html', error='课程名称不能超过100个字符', user=user)
        
        new_course = Course(
            name=name,
            code=code,
            description=description,
            user_id=user.id
        )
        
        db.session.add(new_course)
        db.session.commit()
        
        return redirect(url_for('courses'))
    
    return render_template('add_course.html', user=user)

@app.route('/courses/edit/<int:course_id>', methods=['GET', 'POST'])
def edit_course(course_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    course = Course.query.filter_by(id=course_id, user_id=user.id).first()
    
    if not course:
        return redirect(url_for('courses'))
    
    if request.method == 'POST':
        course.name = request.form.get('name')
        course.code = request.form.get('code')
        course.description = request.form.get('description')
        db.session.commit()
        return redirect(url_for('courses'))
    
    return render_template('edit_course.html', course=course, user=user)

@app.route('/courses/delete/<int:course_id>', methods=['POST'])
def delete_course(course_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'error': '请重新登录'})
    
    course = Course.query.filter_by(id=course_id, user_id=user.id).first()
    
    if course:
        # 清理关联的待确认签到请求
        PendingAttendance.query.filter_by(course_id=course_id, user_id=user.id).delete()
        # 清理关联的签到记录
        Attendance.query.filter_by(course_id=course_id, user_id=user.id).delete()
        db.session.delete(course)
        db.session.commit()
    
    return jsonify({'success': True, 'message': '课程已删除'})

@app.route('/courses/assign/<int:course_id>', methods=['GET', 'POST'])
def assign_classes(course_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    course = Course.query.filter_by(id=course_id, user_id=user.id).first()
    
    if not course:
        return redirect(url_for('courses'))
    
    all_classes = Class.query.filter_by(user_id=user.id).all()
    assigned_class_ids = [cls.id for cls in course.classes]
    
    if request.method == 'POST':
        selected_classes = request.form.getlist('classes')
        course.classes = []
        
        for class_id in selected_classes:
            cls = Class.query.get(class_id)
            if cls and cls.user_id == user.id:
                course.classes.append(cls)
        
        db.session.commit()
        return redirect(url_for('courses'))
    
    return render_template('assign_classes.html', course=course, all_classes=all_classes, assigned_class_ids=assigned_class_ids, user=user)

@app.route('/classes')
def classes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    return render_template('classes.html', classes=classes, user=user)

@app.route('/classes/add', methods=['GET', 'POST'])
def add_class():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        grade = request.form.get('grade', '').strip()
        description = request.form.get('description', '').strip()
        
        if not name:
            return render_template('add_class.html', error='请输入班级名称', user=user)
        
        if len(name) > 100:
            return render_template('add_class.html', error='班级名称不能超过100个字符', user=user)
        
        new_class = Class(
            name=name,
            grade=grade,
            description=description,
            user_id=user.id
        )
        db.session.add(new_class)
        db.session.commit()
        
        return redirect(url_for('classes'))
    
    return render_template('add_class.html', user=user)

@app.route('/classes/edit/<int:class_id>', methods=['GET', 'POST'])
def edit_class(class_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    cls = Class.query.filter_by(id=class_id, user_id=user.id).first_or_404()
    
    if request.method == 'POST':
        cls.name = request.form.get('name')
        cls.grade = request.form.get('grade')
        cls.description = request.form.get('description')
        db.session.commit()
        return redirect(url_for('classes'))
    
    return render_template('edit_class.html', cls=cls, user=user)

@app.route('/classes/delete/<int:class_id>', methods=['POST'])
def delete_class(class_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'error': '请重新登录'})
    
    cls = Class.query.filter_by(id=class_id, user_id=user.id).first_or_404()
    Student.query.filter_by(class_id=class_id, user_id=user.id).delete()
    Attendance.query.filter_by(class_id=class_id, user_id=user.id).delete()
    PendingAttendance.query.filter_by(class_id=class_id, user_id=user.id).delete()
    db.session.delete(cls)
    db.session.commit()
    return jsonify({'success': True, 'message': '班级已删除'})

@app.route('/students')
def students():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    class_id = request.args.get('class_id', type=int)
    
    if class_id:
        students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
        selected_class = Class.query.filter_by(id=class_id, user_id=user.id).first()
    else:
        students = Student.query.filter_by(user_id=user.id).all()
        selected_class = None
    
    return render_template('students.html', 
                         classes=classes, 
                         students=students,
                         selected_class=selected_class,
                         user=user)

@app.route('/students/add', methods=['GET', 'POST'])
def add_student():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        student_id = request.form.get('student_id', '').strip()
        class_id = request.form.get('class_id', type=int)
        
        if not name or not student_id or not class_id:
            return render_template('add_student.html', classes=classes, error='请填写完整信息', user=user)
        
        if len(name) > 100:
            return render_template('add_student.html', classes=classes, error='姓名不能超过100个字符', user=user)
        
        if len(student_id) > 50:
            return render_template('add_student.html', classes=classes, error='学号不能超过50个字符', user=user)
        
        cls = Class.query.filter_by(id=class_id, user_id=user.id).first()
        if not cls:
            return render_template('add_student.html', classes=classes, error='班级不存在', user=user)
        
        new_student = Student(
            name=name,
            student_id=student_id,
            class_id=class_id,
            user_id=user.id
        )
        db.session.add(new_student)
        db.session.commit()
        
        return redirect(url_for('students'))
    
    return render_template('add_student.html', classes=classes, user=user)

@app.route('/students/edit/<int:student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    student = Student.query.filter_by(id=student_id, user_id=user.id).first_or_404()
    classes = Class.query.filter_by(user_id=user.id).all()
    
    if request.method == 'POST':
        student.name = request.form.get('name')
        student.student_id = request.form.get('student_id')
        student.class_id = request.form.get('class_id')
        db.session.commit()
        return redirect(url_for('students'))
    
    return render_template('edit_student.html', student=student, classes=classes, user=user)

@app.route('/students/delete/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'error': '请重新登录'})
    
    student = Student.query.filter_by(id=student_id, user_id=user.id).first_or_404()
    Attendance.query.filter_by(student_id=student_id, user_id=user.id).delete()
    PendingAttendance.query.filter_by(student_id=student_id, user_id=user.id).delete()
    db.session.delete(student)
    db.session.commit()
    return jsonify({'success': True, 'message': '学生已删除'})

@app.route('/capture/<int:student_id>')
def capture(student_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    student = Student.query.filter_by(id=student_id, user_id=user.id).first_or_404()
    cls = Class.query.filter_by(id=student.class_id, user_id=user.id).first()
    return render_template('capture_face.html', student=student, cls=cls, user=user)

@app.route('/api/save_descriptor/<int:student_id>', methods=['POST'])
def save_descriptor(student_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        if not user:
            return jsonify({'success': False, 'error': '请重新登录'})
        
        student = Student.query.filter_by(id=student_id, user_id=user.id).first_or_404()
        
        data = request.json
        if not data or 'descriptor' not in data:
            return jsonify({'success': False, 'error': '缺少人脸特征数据'})
        
        descriptor = np.array(data['descriptor'], dtype=np.float32)
        student.face_descriptor = descriptor.tobytes()
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': '保存人脸特征失败'})

@app.route('/api/save_face', methods=['POST'])
def save_face():
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '未登录'})
        
        user = get_current_user()
        if not user:
            return jsonify({'success': False, 'message': '请重新登录'})
        
        data = request.json
        if not data or 'student_id' not in data or 'descriptor' not in data:
            return jsonify({'success': False, 'message': '缺少必要参数'})
        
        student_id = data['student_id']
        descriptor = np.array(data['descriptor'], dtype=np.float32)
        descriptor = descriptor / (np.linalg.norm(descriptor) + 1e-8)
        image_b64 = data.get('image')
        
        student = Student.query.filter_by(id=student_id, user_id=user.id).first()
        
        if not student:
            return jsonify({'success': False, 'message': '学生不存在'})
        
        student.face_descriptor = descriptor.tobytes()
        
        # 同时采集 512 维特征（InsightFace）
        if image_b64:
            try:
                recognizer = get_recognizer()
                if recognizer.use_real_models:
                    img = recognizer.base64_to_numpy(image_b64)
                    faces = recognizer.detect_faces(img)
                    if faces:
                        descriptor_512 = faces[0].get('embedding')
                        if descriptor_512 is not None:
                            student.face_descriptor_512 = descriptor_512.astype(np.float32).tobytes()
                            print(f'[SaveFace] 512 维特征已保存: {len(descriptor_512)} 维')
                        else:
                            print('[SaveFace] 警告: 未提取到 512 维特征')
                    else:
                        print('[SaveFace] 警告: 未检测到人脸，512 维特征未保存')
                else:
                    print('[SaveFace] InsightFace 未启用，跳过 512 维特征采集')
            except Exception as e:
                print(f'[SaveFace] 512 维特征采集失败: {e}')
        
        db.session.commit()
        has_512 = student.face_descriptor_512 is not None
        msg = '人脸采集成功（128维+512维双特征）' if has_512 else '人脸采集成功（仅128维，后端识别不可用）'
        print(f'[SaveFace] {msg}')
        return jsonify({'success': True, 'message': msg, 'has_512': has_512})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '人脸采集失败'})

@app.route('/attendance')
def attendance():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    courses = Course.query.filter_by(user_id=user.id).all()
    return render_template('attendance.html', classes=classes, courses=courses, user=user)

@app.route('/batch_attendance')
def batch_attendance():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    courses = Course.query.filter_by(user_id=user.id).all()
    return render_template('batch_attendance.html', courses=courses, user=user)

@app.route('/api/get_courses')
def get_courses():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    courses = Course.query.filter_by(user_id=user.id).all()
    
    course_list = []
    for course in courses:
        course_list.append({
            'id': course.id,
            'name': course.name,
            'code': course.code,
            'description': course.description,
            'class_count': len(course.classes)
        })
    
    return jsonify({'success': True, 'courses': course_list})

@app.route('/api/get_course_info/<int:course_id>')
def get_course_info(course_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        course = Course.query.filter_by(id=course_id, user_id=user.id).first()
        
        if not course:
            return jsonify({'success': False, 'error': '课程不存在'})
        
        classes = course.classes
        class_ids = [cls.id for cls in classes]
        
        students = Student.query.filter(Student.class_id.in_(class_ids), Student.user_id == user.id).all()
        
        today = date.today()
        student_data = []
        
        for student in students:
            attendance = Attendance.query.filter(
                Attendance.student_id == student.id,
                Attendance.course_id == course_id,
                db.func.date(Attendance.check_in_time) == today
            ).first()
            
            face_desc = None
            if student.face_descriptor:
                face_desc = base64.b64encode(student.face_descriptor).decode('utf-8')
            
            student_data.append({
                'id': student.id,
                'name': student.name,
                'student_id': student.student_id,
                'class_id': student.class_id,
                'class_name': next((c.name for c in classes if c.id == student.class_id), '未知'),
                'face_descriptor': face_desc,
                'checked_in': attendance is not None,
                'check_in_time': attendance.check_in_time.isoformat() if attendance else None
            })
        
        class_data = [{
            'id': cls.id,
            'name': cls.name,
            'grade': cls.grade
        } for cls in classes]
        
        return jsonify({
            'success': True,
            'course_name': course.name,
            'course_code': course.code,
            'classes': class_data,
            'students': student_data
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': '获取课程信息失败'})

@app.route('/api/batch_checkin', methods=['POST'])
def batch_checkin():
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '未登录'})
        
        user = get_current_user()
        data = request.json
        student_ids = data.get('student_ids', [])
        course_id = data.get('course_id')
        
        if not student_ids or not course_id:
            return jsonify({'success': False, 'message': '参数不足'})
        
        today = date.today()
        count = 0
        
        for student_id in student_ids:
            student = Student.query.filter_by(id=student_id, user_id=user.id).first()
            if not student:
                continue
            
            # 统一重复签到检查：同一学生同一天只能签到一次
            existing = Attendance.query.filter(
                Attendance.student_id == student_id,
                Attendance.user_id == user.id,
                db.func.date(Attendance.check_in_time) == today
            ).first()
            
            if existing:
                continue
            
            attendance = Attendance(
                student_id=student_id,
                class_id=student.class_id,
                course_id=course_id,
                user_id=user.id,
                method='face'
            )
            db.session.add(attendance)
            count += 1
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'成功签到 {count} 人'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '批量签到失败'})

@app.route('/api/get_students/<int:class_id>')
def get_students(class_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
        today = date.today()
        
        result = []
        for student in students:
            attended = Attendance.query.filter(
                Attendance.student_id == student.id,
                Attendance.user_id == user.id,
                db.func.date(Attendance.check_in_time) == today
            ).first() is not None
            
            result.append({
                'id': student.id,
                'name': student.name,
                'student_id': student.student_id,
                'face_descriptor': np.frombuffer(student.face_descriptor, dtype=np.float32).tolist() if student.face_descriptor else None,
                'attended': attended
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Error in get_students: {e}")
        return jsonify([])

@app.route('/api/record_attendance', methods=['POST'])
def record_attendance():
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        if not user:
            return jsonify({'success': False, 'error': '请重新登录'})
        
        data = request.json
        if not data or 'student_id' not in data or 'class_id' not in data:
            return jsonify({'success': False, 'error': '参数不足'})
        
        student_id = data['student_id']
        class_id = data['class_id']
        course_id = data.get('course_id')
        method = data.get('method', 'face')
        
        student = Student.query.filter_by(id=student_id, user_id=user.id).first()
        if not student:
            return jsonify({'success': False, 'message': '学生不存在'})
        
        today = date.today()
        existing = Attendance.query.filter(
            Attendance.student_id == student_id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == today
        ).first()
        
        if existing:
            return jsonify({'success': False, 'already_attended': True})
        
        attendance = Attendance(
            student_id=student_id,
            class_id=class_id,
            course_id=course_id,
            method=method,
            user_id=user.id
        )
        db.session.add(attendance)
        db.session.commit()
        
        return jsonify({'success': True, 'student_name': student.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '签到失败'})

@app.route('/api/clear_attendance/<int:class_id>', methods=['POST'])
def clear_attendance(class_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    if not user:
        return jsonify({'success': False, 'error': '请重新登录'})
    today = date.today()
    Attendance.query.filter(
        Attendance.class_id == class_id,
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).delete()
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/clear_today_attendance/<int:class_id>', methods=['POST'])
def clear_today_attendance(class_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    today = date.today()
    Attendance.query.filter(
        Attendance.class_id == class_id,
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': '今日签到数据已清除'})

@app.route('/api/clear_today_attendance_by_course/<int:course_id>', methods=['POST'])
def clear_today_attendance_by_course(course_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    today = date.today()
    
    course = Course.query.filter_by(id=course_id, user_id=user.id).first()
    if not course:
        return jsonify({'success': False, 'error': '课程不存在'})
    
    Attendance.query.filter(
        Attendance.course_id == course_id,
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': '课程今日签到数据已清除'})

from face_recognition_backend import get_recognizer, get_liveness_tracker


@app.route('/api/recognize_face', methods=['POST'])
def recognize_face():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    data = request.json
    class_id = data['class_id']
    course_id = data.get('course_id')
    descriptor = np.array(data['descriptor'], dtype=np.float32)
    descriptor = descriptor / (np.linalg.norm(descriptor) + 1e-8)
    
    students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
    
    min_distance = float('inf')
    matched_student = None
    
    for student in students:
        if student.face_descriptor:
            stored_descriptor = np.frombuffer(student.face_descriptor, dtype=np.float32)
            stored_descriptor = stored_descriptor / (np.linalg.norm(stored_descriptor) + 1e-8)
            distance = np.linalg.norm(descriptor - stored_descriptor)
            print(f'[RecognizeFace] {student.name} distance={distance:.4f}')
            if distance < min_distance and distance < 0.6:
                min_distance = distance
                matched_student = student
    
    if matched_student:
        print(f'[RecognizeFace] 匹配成功: {matched_student.name} (distance={min_distance:.4f})')
        today = date.today()
        existing = Attendance.query.filter(
            Attendance.student_id == matched_student.id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == today
        ).first()
        
        if existing:
            return jsonify({'success': True, 'message': matched_student.name + ' 今日已签到', 'student_id': matched_student.id})
        
        attendance = Attendance(
            student_id=matched_student.id,
            class_id=class_id,
            course_id=course_id,
            method='face',
            user_id=user.id
        )
        db.session.add(attendance)
        db.session.commit()
        
        return jsonify({'success': True, 'message': matched_student.name + ' 签到成功', 'student_id': matched_student.id})
    else:
        return jsonify({'success': False, 'message': '未识别到匹配的学生'})


@app.route('/api/recognize_frame', methods=['POST'])
def recognize_frame():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    data = request.json
    class_id = data.get('class_id')
    course_id = data.get('course_id')
    image_b64 = data.get('image')
    
    if not image_b64:
        return jsonify({'success': False, 'message': '缺少图像数据'})
    
    # 加载班级学生特征库（后端模式使用 512 维）
    students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
    students_dict = {
        s.id: s.face_descriptor_512
        for s in students if s.face_descriptor_512
    }
    
    if not students_dict:
        return jsonify({'success': False, 'message': '班级无已录入人脸的学生（需重新采集人脸以启用后端识别）'})
    
    # 后端识别
    recognizer = get_recognizer()
    result = recognizer.recognize(image_b64, students_dict)
    recognized = result['recognized']
    face_count = result['detected_faces']
    liveness_ok = result.get('liveness', False)
    nose_frames = result.get('nose_frames', 0)
    
    print(f'[RecognizeFrame] 检测到 {face_count} 张人脸, 匹配 {len(recognized)} 人, 活体={liveness_ok} ({nose_frames}帧)')
    
    if len(recognized) == 0:
        if face_count > 0 and not liveness_ok:
            msg = f'请轻微移动头部（已采集 {nose_frames} 帧，需 6 帧）' if nose_frames > 0 else '请正对摄像头并轻微移动头部'
        else:
            msg = '未检测到人脸，请正对摄像头' if face_count == 0 else f'检测到 {face_count} 张人脸，但未匹配到学生'
        return jsonify({
            'success': False,
            'message': msg,
            'face_count': face_count,
            'liveness': liveness_ok,
            'nose_frames': nose_frames,
            'matched': []
        })
    
    # 处理第一个匹配（多人场景后续可扩展）
    match = recognized[0]
    matched_student = Student.query.filter_by(id=match['student_id'], user_id=user.id).first()
    
    if not matched_student:
        return jsonify({'success': False, 'message': '学生不存在'})
    
    today = date.today()
    existing = Attendance.query.filter(
        Attendance.student_id == matched_student.id,
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).first()
    
    if existing:
        return jsonify({
            'success': True,
            'already_attended': True,
            'message': f'{matched_student.name} 今日已签到',
            'student_id': matched_student.id,
            'student_name': matched_student.name,
            'matched': [{'student_id': m['student_id'], 'bbox': m['bbox']} for m in recognized]
        })
    
    attendance = Attendance(
        student_id=matched_student.id,
        class_id=class_id,
        course_id=course_id,
        method='face',
        user_id=user.id
    )
    db.session.add(attendance)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'{matched_student.name} 签到成功',
        'student_id': matched_student.id,
        'student_name': matched_student.name,
        'liveness': liveness_ok,
        'nose_frames': nose_frames,
        'matched': [{'student_id': m['student_id'], 'bbox': m['bbox']} for m in recognized]
    })

@app.route('/api/reset_liveness', methods=['POST'])
def reset_liveness():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    get_liveness_tracker().reset()
    print('[Liveness] 活体追踪器已重置')
    return jsonify({'success': True})

@app.route('/api/checkin_manual', methods=['POST'])
def checkin_manual():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': '未登录'})
    
    user = get_current_user()
    data = request.json
    student_id = data['student_id']
    class_id = data['class_id']
    course_id = data.get('course_id')
    
    student = Student.query.filter_by(id=student_id, user_id=user.id).first()
    
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'})
    
    today = date.today()
    # 统一重复签到检查：同一学生同一天只能签到一次
    existing = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.user_id == user.id,
        db.func.date(Attendance.check_in_time) == today
    ).first()
    
    if existing:
        return jsonify({'success': False, 'message': student.name + ' 今日已签到'})
    
    attendance = Attendance(
        student_id=student_id,
        class_id=class_id,
        course_id=course_id,
        method='manual',
        user_id=user.id
    )
    db.session.add(attendance)
    db.session.commit()
    
    return jsonify({'success': True, 'message': student.name + ' 手动签到成功'})

@app.route('/api/get_records', methods=['POST'])
def get_records():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'})
    
    user = get_current_user()
    data = request.json
    class_id = data.get('class_id')
    selected_date = data.get('date')
    
    query = Attendance.query.filter_by(user_id=user.id)
    
    if class_id:
        query = query.filter_by(class_id=class_id)
    
    if selected_date:
        query = query.filter(db.func.date(Attendance.check_in_time) == selected_date)
    
    attendances = query.order_by(Attendance.check_in_time.desc()).all()
    
    result = []
    for a in attendances:
        student = Student.query.get(a.student_id)
        cls = Class.query.get(a.class_id)
        result.append({
            'id': a.id,
            'student_name': student.name if student else '未知',
            'student_id': student.student_id if student else '',
            'class_name': cls.name if cls else '未知',
            'check_in_time': a.check_in_time.strftime('%Y-%m-%d %H:%M:%S'),
            'method': a.method
        })
    
    return jsonify(result)

@app.route('/records')
def records():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    class_id = request.args.get('class_id', type=int)
    selected_date = request.args.get('date')
    
    if class_id and selected_date:
        attendances = Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == selected_date
        ).order_by(Attendance.check_in_time.desc()).all()
        selected_class = Class.query.filter_by(id=class_id, user_id=user.id).first()
    elif class_id:
        attendances = Attendance.query.filter_by(class_id=class_id, user_id=user.id).order_by(Attendance.check_in_time.desc()).all()
        selected_class = Class.query.filter_by(id=class_id, user_id=user.id).first()
    else:
        attendances = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.check_in_time.desc()).all()
        selected_class = None
    
    return render_template('records.html',
                         classes=classes,
                         attendances=attendances,
                         selected_class=selected_class,
                         selected_date=selected_date,
                         user=user)

@app.route('/statistics')
def statistics():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    classes = Class.query.filter_by(user_id=user.id).all()
    class_id = request.args.get('class_id', type=int)
    
    if class_id:
        selected_class = Class.query.filter_by(id=class_id, user_id=user.id).first()
        students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
        total_students = len(students)
        
        today = date.today()
        week_ago = today - timedelta(days=7)
        daily_stats = []
        
        for i in range(7):
            day = week_ago + timedelta(days=i)
            count = Attendance.query.filter(
                Attendance.class_id == class_id,
                Attendance.user_id == user.id,
                db.func.date(Attendance.check_in_time) == day
            ).count()
            daily_stats.append({
                'date': day.strftime('%Y-%m-%d'),
                'count': count
            })
        
        hour_stats = [0] * 24
        attendances = Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.user_id == user.id,
            Attendance.check_in_time >= week_ago
        ).all()
        for a in attendances:
            hour_stats[a.check_in_time.hour] += 1
        
        method_stats = {}
        for a in attendances:
            method_stats[a.method] = method_stats.get(a.method, 0) + 1
        
        student_ranking = []
        for student in students:
            count = Attendance.query.filter(
                Attendance.student_id == student.id,
                Attendance.user_id == user.id,
                Attendance.check_in_time >= week_ago
            ).count()
            student_ranking.append({
                'name': student.name,
                'student_id': student.student_id,
                'count': count
            })
        student_ranking.sort(key=lambda x: x['count'], reverse=True)
    else:
        selected_class = None
        daily_stats = []
        hour_stats = []
        method_stats = {}
        student_ranking = []
        total_students = 0
    
    return render_template('statistics.html',
                         classes=classes,
                         selected_class=selected_class,
                         daily_stats=daily_stats,
                         hour_stats=hour_stats,
                         method_stats=method_stats,
                         student_ranking=student_ranking,
                         total_students=total_students,
                         user=user)

@app.route('/api/export_statistics')
def export_statistics():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "总体统计"
    
    today = date.today()
    week_ago = today - timedelta(days=7)
    
    ws1.append(['日期', '签到人数'])
    for i in range(7):
        day = week_ago + timedelta(days=i)
        count = Attendance.query.filter(
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == day
        ).count()
        ws1.append([day.strftime('%Y-%m-%d'), count])
    
    ws2 = wb.create_sheet(title="班级统计")
    ws2.append(['班级', '学生数', '近7天签到次数'])
    classes = Class.query.filter_by(user_id=user.id).all()
    for cls in classes:
        student_count = Student.query.filter_by(class_id=cls.id, user_id=user.id).count()
        attendance_count = Attendance.query.filter(
            Attendance.class_id == cls.id,
            Attendance.user_id == user.id,
            Attendance.check_in_time >= week_ago
        ).count()
        ws2.append([cls.name, student_count, attendance_count])
    
    ws3 = wb.create_sheet(title="详细记录")
    ws3.append(['姓名', '学号', '班级', '签到时间', '方式'])
    attendances = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.check_in_time.desc()).limit(100).all()
    for a in attendances:
        student = Student.query.get(a.student_id)
        cls = Class.query.get(a.class_id)
        ws3.append([student.name, student.student_id, cls.name, a.check_in_time.strftime('%Y-%m-%d %H:%M:%S'), a.method])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name='签到统计.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export_classes')
def export_classes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    wb = Workbook()
    ws = wb.active
    ws.title = "班级信息"
    ws.append(['ID', '班级名称', '年级', '描述'])
    classes = Class.query.filter_by(user_id=user.id).all()
    for cls in classes:
        ws.append([cls.id, cls.name, cls.grade or '', cls.description or ''])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='班级信息.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export_students/<int:class_id>')
def export_students(class_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    ws.append(['ID', '姓名', '学号', '班级'])
    if class_id == 0:
        students = Student.query.filter_by(user_id=user.id).all()
    else:
        students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
    for student in students:
        cls = Class.query.get(student.class_id)
        ws.append([student.id, student.name, student.student_id, cls.name])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='学生信息.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/import_classes', methods=['POST'])
def import_classes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    if 'file' not in request.files:
        return redirect(url_for('classes'))
    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('classes'))
    if file:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                existing = Class.query.filter_by(name=row[1], user_id=user.id).first()
                if not existing:
                    new_class = Class(
                        name=row[1],
                        grade=row[2] if len(row) > 2 else None,
                        description=row[3] if len(row) > 3 else None,
                        user_id=user.id
                    )
                    db.session.add(new_class)
        db.session.commit()
    return redirect(url_for('classes'))

@app.route('/api/import_students', methods=['POST'])
def import_students():
    if 'user_id' not in session:
        return redirect(url_for('students'))
    
    user = get_current_user()
    class_id = request.form.get('class_id')
    if 'file' not in request.files:
        return redirect(url_for('students'))
    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('students'))
    if file and class_id:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:
                existing = Student.query.filter_by(student_id=row[2], user_id=user.id).first()
                if not existing:
                    new_student = Student(
                        name=row[1],
                        student_id=row[2],
                        class_id=class_id,
                        user_id=user.id
                    )
                    db.session.add(new_student)
        db.session.commit()
    return redirect(url_for('students'))

@app.route('/super_admin')
def super_admin_dashboard():
    if session.get('role') != 'super_admin':
        return redirect(url_for('login'))
    
    all_users = User.query.all()
    
    total_users = len(all_users)
    total_classes = 0
    total_students = 0
    total_attendances = 0
    today_total_attendance = 0
    
    all_attendances = []
    all_classes = []
    
    for user in all_users:
        total_classes += len(user.classes)
        total_students += len(user.students)
        total_attendances += len(user.attendances)
        all_attendances.extend(user.attendances)
        all_classes.extend(user.classes)
    
    today = datetime.now().date()
    today_attendances = [a for a in all_attendances if a.check_in_time and a.check_in_time.date() == today]
    today_total_attendance = len(today_attendances)
    
    avg_students_per_class = total_students / total_classes if total_classes > 0 else 0
    
    from collections import defaultdict
    
    user_stats = []
    for user in all_users:
        user_attendances = Attendance.query.filter_by(user_id=user.id).all()
        today_user_attendances = [a for a in user_attendances if a.check_in_time and a.check_in_time.date() == today]
        
        student_attendance_rate = 0
        if len(user.students) > 0:
            attended_students = len(set([a.student_id for a in today_user_attendances]))
            student_attendance_rate = (attended_students / len(user.students)) * 100
        
        user_stats.append({
            'id': user.id,
            'username': user.username,
            'name': user.name,
            'class_count': len(user.classes),
            'student_count': len(user.students),
            'attendance_count': len(user_attendances),
            'today_attendance': len(today_user_attendances),
            'attendance_rate': round(student_attendance_rate, 1)
        })
    
    class_stats = []
    for cls in all_classes:
        class_attendances = Attendance.query.filter_by(class_id=cls.id).all()
        today_class_attendances = [a for a in class_attendances if a.check_in_time and a.check_in_time.date() == today]
        
        class_stats.append({
            'id': cls.id,
            'name': cls.name,
            'grade': cls.grade,
            'student_count': len(cls.students),
            'total_attendances': len(class_attendances),
            'today_attendance': len(today_class_attendances),
            'admin_name': cls.user.name if cls.user else '未知'
        })
    
    date_attendance = defaultdict(int)
    for attendance in all_attendances:
        if attendance.check_in_time:
            date_str = attendance.check_in_time.strftime('%Y-%m-%d')
            date_attendance[date_str] += 1
    
    date_list = sorted(date_attendance.keys())[-7:]
    date_data = [{'date': d, 'count': date_attendance[d]} for d in date_list]
    
    max_count = max((item['count'] for item in date_data), default=1)
    
    return render_template('super_admin.html', 
                          total_users=total_users,
                          total_classes=total_classes,
                          total_students=total_students,
                          total_attendances=total_attendances,
                          today_total_attendance=today_total_attendance,
                          avg_students_per_class=round(avg_students_per_class, 1),
                          user_stats=user_stats,
                          class_stats=class_stats,
                          date_data=date_data,
                          max_count=max_count)

@app.route('/api/get_qrcode/<token>')
def get_qrcode(token):
    try:
        qr_url = url_for('qr_signin', token=token, _external=True)
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color='black', back_color='white')
        
        img_buffer = BytesIO()
        img.save(img_buffer)
        img_buffer.seek(0)
        
        return send_file(img_buffer, mimetype='image/png')
    except Exception as e:
        return jsonify({'success': False, 'error': '生成二维码失败'})

@app.route('/qrcode_attendance')
def qrcode_attendance():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = get_current_user()
    courses = Course.query.filter_by(user_id=user.id).all()
    return render_template('qrcode_attendance.html', courses=courses, user=user)

@app.route('/qr_signin/<token>')
def qr_signin(token):
    parts = token.split('_')
    if len(parts) < 2:
        return render_template('qr_signin_error.html', error='无效的签到链接')
    
    course_id = parts[0]
    try:
        course_id = int(course_id)
    except ValueError:
        return render_template('qr_signin_error.html', error='无效的课程ID')
    
    user = None
    if 'user_id' in session:
        user = get_current_user()
    
    course = Course.query.filter_by(id=course_id).first()
    if not course:
        return render_template('qr_signin_error.html', error='课程不存在')
    
    user_id = course.user_id
    return render_template('qr_signin.html', course=course, token=token, user_id=user_id)

@app.route('/api/qr_checkin', methods=['POST'])
def qr_checkin():
    try:
        data = request.json
        token = data.get('token')
        student_id = data.get('student_id')
        
        if not token or not student_id:
            return jsonify({'success': False, 'message': '参数不足'})
        
        parts = token.split('_')
        if len(parts) < 2:
            return jsonify({'success': False, 'message': '无效的签到链接'})
        
        course_id = parts[0]
        try:
            course_id = int(course_id)
        except ValueError:
            return jsonify({'success': False, 'message': '无效的课程ID'})
        
        course = Course.query.filter_by(id=course_id).first()
        if not course:
            return jsonify({'success': False, 'message': '课程不存在'})
        
        user_id = course.user_id
        
        student = Student.query.filter_by(student_id=student_id, user_id=user_id).first()
        if not student:
            return jsonify({'success': False, 'message': '学号不存在或不属于该课程'})
        
        course_class_ids = [cls.id for cls in course.classes]
        if student.class_id not in course_class_ids:
            return jsonify({'success': False, 'message': '该学生不属于本课程'})
        
        today = date.today()
        existing = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.user_id == user_id,
            db.func.date(Attendance.check_in_time) == today
        ).first()
        
        if existing:
            return jsonify({'success': False, 'message': f'{student.name} 今日已签到'})
        
        pending = PendingAttendance.query.filter(
            PendingAttendance.student_id == student.id,
            PendingAttendance.course_id == course_id,
            PendingAttendance.status == 'pending',
            PendingAttendance.request_time >= datetime.now() - timedelta(minutes=5)
        ).first()
        
        if pending:
            return jsonify({'success': False, 'message': '签到请求已发送，请等待教师确认'})
        
        pending_attendance = PendingAttendance(
            student_id=student.id,
            class_id=student.class_id,
            course_id=course_id,
            user_id=user_id,
            method='qrcode'
        )
        db.session.add(pending_attendance)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'{student.name} 签到请求已发送，请等待教师确认', 'needs_confirm': True})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '二维码签到请求失败'})

@app.route('/api/get_pending_attendances/<int:course_id>')
def get_pending_attendances(course_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        
        pending_list = PendingAttendance.query.filter(
            PendingAttendance.course_id == course_id,
            PendingAttendance.user_id == user.id,
            PendingAttendance.status == 'pending',
            PendingAttendance.request_time >= datetime.now() - timedelta(minutes=5)
        ).all()
        
        pending_data = []
        for pending in pending_list:
            student = Student.query.get(pending.student_id)
            if not student:
                continue
            pending_data.append({
                'id': pending.id,
                'student_id': student.student_id,
                'name': student.name,
                'request_time': pending.request_time.isoformat()
            })
        
        return jsonify({'success': True, 'pending': pending_data})
    except Exception as e:
        return jsonify({'success': False, 'error': '获取待确认签到失败'})

@app.route('/api/confirm_attendance/<int:pending_id>', methods=['POST'])
def confirm_attendance(pending_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        
        pending = PendingAttendance.query.filter_by(id=pending_id, user_id=user.id).first()
        if not pending:
            return jsonify({'success': False, 'message': '签到请求不存在'})
        
        if pending.status != 'pending':
            return jsonify({'success': False, 'message': '签到请求已处理'})
        
        today = date.today()
        # 统一重复签到检查：同一学生同一天只能签到一次
        existing = Attendance.query.filter(
            Attendance.student_id == pending.student_id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == today
        ).first()
        
        if existing:
            pending.status = 'rejected'
            db.session.commit()
            student = Student.query.get(pending.student_id)
            return jsonify({'success': False, 'message': student.name + ' 今日已签到'})
        
        attendance = Attendance(
            student_id=pending.student_id,
            class_id=pending.class_id,
            course_id=pending.course_id,
            user_id=pending.user_id,
            method=pending.method
        )
        db.session.add(attendance)
        
        pending.status = 'confirmed'
        db.session.commit()
        
        return jsonify({'success': True, 'message': '确认签到成功'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '确认签到失败'})

@app.route('/api/reject_attendance/<int:pending_id>', methods=['POST'])
def reject_attendance(pending_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        
        pending = PendingAttendance.query.filter_by(id=pending_id, user_id=user.id).first()
        if not pending:
            return jsonify({'success': False, 'message': '签到请求不存在'})
        
        pending.status = 'rejected'
        db.session.commit()
        
        return jsonify({'success': True, 'message': '已拒绝该签到请求'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': '拒绝签到失败'})

@app.route('/api/get_statistics/<int:class_id>')
def get_statistics(class_id):
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '未登录'})
        
        user = get_current_user()
        selected_class = Class.query.filter_by(id=class_id, user_id=user.id).first()
        
        if not selected_class:
            return jsonify({'success': False, 'error': '班级不存在'})
        
        students = Student.query.filter_by(class_id=class_id, user_id=user.id).all()
        total_students = len(students)
        
        today = date.today()
        week_ago = today - timedelta(days=7)
        
        # 近7天每日数据
        weekly_data = []
        for i in range(7):
            day = week_ago + timedelta(days=i)
            count = Attendance.query.filter(
                Attendance.class_id == class_id,
                Attendance.user_id == user.id,
                db.func.date(Attendance.check_in_time) == day
            ).count()
            weekly_data.append({
                'date': day.strftime('%Y-%m-%d'),
                'count': count
            })
        
        # 今日签到
        today_attendance = Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.user_id == user.id,
            db.func.date(Attendance.check_in_time) == today
        ).count()
        
        today_rate = (today_attendance / total_students) * 100 if total_students > 0 else 0
        
        # 签到时间分布
        hour_distribution = [0] * 24
        attendances_week = Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.user_id == user.id,
            Attendance.check_in_time >= week_ago
        ).all()
        for a in attendances_week:
            hour_distribution[a.check_in_time.hour] += 1
        
        # 签到方式分布
        method_distribution = {'face': 0, 'manual': 0, 'qrcode': 0}
        for a in attendances_week:
            if a.method in method_distribution:
                method_distribution[a.method] += 1
        
        # 总签到次数
        total_attendances = len(attendances_week)
        
        # 人均签到次数
        avg_attendance = total_attendances / total_students if total_students > 0 else 0
        
        # 学生出勤排名
        student_attendance = []
        for student in students:
            count = Attendance.query.filter(
                Attendance.student_id == student.id,
                Attendance.user_id == user.id,
                Attendance.check_in_time >= week_ago
            ).count()
            student_attendance.append({
                'name': student.name,
                'count': count
            })
        student_attendance.sort(key=lambda x: x['count'], reverse=True)
        
        return jsonify({
            'success': True,
            'total_students': total_students,
            'today_attendance': today_attendance,
            'today_rate': today_rate,
            'total_attendances': total_attendances,
            'avg_attendance': avg_attendance,
            'weekly_data': weekly_data,
            'hour_distribution': hour_distribution,
            'method_distribution': method_distribution,
            'student_attendance': student_attendance
        })
    except Exception as e:
        return jsonify({'success': False, 'error': '获取统计数据失败'})

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
